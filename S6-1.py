# ----------------------------------------------------------------------
# This Program is for making a Final Dataset out of tweets that all are judged
# ----------------------------------------------------------------------
def check_if_dataset_by_judge_is_correct (worksheet1_whole):
    Tweet_State = worksheet1_whole.col_values(6)
    del Tweet_State[0]
    number_tweet = len (Tweet_State)
    k = 0
    vote = []
    while k < number_tweet :
        vote.append (Tweet_State[k].lower())
        k += 1
        
    i = 0
    accepted_letters = ["r","a","q","n"]
    x = "Dataset is Ok"
    while i < number_tweet :
        if vote[i] not in accepted_letters :
            print ("Non Standard Letter in Amir's votes is :", vote[i])
            print ("Non Standard Letter in Amir's votes is in Row Number :", i + 2)
            x = "Dataset Needs to be corrected"

        i += 1

    if x != "Dataset is Ok":
        print (x)
    return x

    
def make_dic_retweet_frequency_state (worksheet1_whole):
    Frequency = worksheet1_whole.col_values(1)
    Retweet_ID = worksheet1_whole.col_values(2)
    Tweet_ID = worksheet1_whole.col_values(0)
    Tweet_State_might_have_captals = worksheet1_whole.col_values(6)
    del Frequency[0]
    del Retweet_ID[0]
    del Tweet_ID[0]
    del Tweet_State_might_have_captals[0]
    
    Tweet_State = []
    for state in Tweet_State_might_have_captals :
        Tweet_State.append (state.lower())
        

    number_tweets = len (Frequency)    
    dic_retweet_frequency_state = {}
    k = 0
    while k < number_tweets :
        if Retweet_ID[k] != 'Not a Retweet' :
            dic_retweet_frequency_state[Retweet_ID[k]]= Frequency[k],Tweet_State[k],"r" #the reason to put this letter "r" here will be figured out in function "make_list_of_states_frequencies"
        else :
            dic_retweet_frequency_state[Tweet_ID[k]]= Frequency[k],Tweet_State[k],"t" #the reason to put this letter "t" here will be figured out in function "make_list_of_states_frequencies"
            
        k += 1



    return dic_retweet_frequency_state



def make_list_of_states_frequencies (worksheet1_main,dic_retweet_frequency_state):
    # first check to see if the tweet is a retweet or non retweet, because if it is a retweet then we want the retweet id as tweet id
    retweet_id = worksheet1_main.col_values(23)
    tweet_id = worksheet1_main.col_values(21)
    del retweet_id[0]
    del tweet_id[0]
    
    tweet_number = len (tweet_id)
    i = 0
    id_list = []
    while i < tweet_number :
        if retweet_id[i] != 'Not a Retweet' :
            id_list.append ([retweet_id[i],"r"])
        else :
            id_list.append ([tweet_id[i],"t"])

        i+=1

    Frequency_List = []
    State_List = []
    for n in id_list :
        for m in dic_retweet_frequency_state :
            if m == n[0] :
                if ((n[1] == "t") and (dic_retweet_frequency_state[m][2] == "r")) :
                    statement = "This Tweet is Origin of " + str(dic_retweet_frequency_state[m][0]) + " Tweets"
                    Frequency_List.append (statement)
                    State_List.append (dic_retweet_frequency_state[m][1])

                else :
                    Frequency_List.append (str(dic_retweet_frequency_state[m][0]))
                    State_List.append (dic_retweet_frequency_state[m][1])



    return Frequency_List,State_List
                


def make_final_dataset (State_List,my_path_input,my_path_output,Frequency_List) : #it is same as

    import openpyxl as xl

    wb1 = xl.load_workbook(my_path_input + 'excel_results.xlsx')
    
    ws1_1 = wb1.worksheets[0]
    ws1_2 = wb1.worksheets[1]
    ws1_3 = wb1.worksheets[2]
    ws1_4 = wb1.worksheets[3]
    ws1_5 = wb1.worksheets[4]
    ws1_6 = wb1.worksheets[5]
    ws1_7 = wb1.worksheets[6]
    ws1_8 = wb1.worksheets[7]
    ws1_9 = wb1.worksheets[8]
    ws1_10 = wb1.worksheets[9]
    ws1_11 = wb1.worksheets[10]
    ws1_12 = wb1.worksheets[11]

    
    wb2 = xl.Workbook()

    # Sheet 1 -------
    ws2_1 = wb2.create_sheet(ws1_1.title)
    for row in ws1_1:
        for cell in row:
            ws2_1[cell.coordinate].value = cell.value


    #writing the frequency of each tweet in column 27
    ws2_1.cell(row = 1 ,column = 28).value = 'Frequency of Tweet Occurance' #for excel 'row = 1' is same as index = 0 for python , ie the label row
    v = 2
    for fr in Frequency_List :
        ws2_1.cell(row = v ,column = 28).value = fr
        v += 1
    
    
    #writing state of each tweet in column 28
    ws2_1.cell(row = 1 ,column = 29).value = 'State of Tweet' #for excel 'row = 1' is same as index = 0 for python , ie the label row
    v = 2
    for state in State_List :
        ws2_1.cell(row = v ,column = 29).value = state
        v += 1

    ws2_1.column_dimensions['A'].width = 25
    ws2_1.column_dimensions['B'].width = 25
    ws2_1.column_dimensions['C'].width = 200
    ws2_1.column_dimensions['D'].width = 15
    ws2_1.column_dimensions['E'].width = 25
    ws2_1.column_dimensions['F'].width = 10
    ws2_1.column_dimensions['G'].width = 30
    ws2_1.column_dimensions['H'].width = 35
    ws2_1.column_dimensions['I'].width = 15
    ws2_1.column_dimensions['J'].width = 15
    ws2_1.column_dimensions['K'].width = 25
    ws2_1.column_dimensions['L'].width = 15
    ws2_1.column_dimensions['M'].width = 25
    ws2_1.column_dimensions['N'].width = 15
    ws2_1.column_dimensions['O'].width = 20
    ws2_1.column_dimensions['P'].width = 20
    ws2_1.column_dimensions['Q'].width = 20
    ws2_1.column_dimensions['R'].width = 15
    ws2_1.column_dimensions['S'].width = 50
    ws2_1.column_dimensions['T'].width = 50
    ws2_1.column_dimensions['U'].width = 200
    ws2_1.column_dimensions['V'].width = 25
    ws2_1.column_dimensions['W'].width = 200
    ws2_1.column_dimensions['X'].width = 25
    ws2_1.column_dimensions['Y'].width = 200
    ws2_1.column_dimensions['Z'].width = 25
    ws2_1.column_dimensions['AA'].width = 25
    ws2_1.column_dimensions['AB'].width = 30
    ws2_1.column_dimensions['AC'].width = 15

    
    
    # Sheet 2 -------
    ws2_2 = wb2.create_sheet(ws1_2.title)
    for row in ws1_2:
        for cell in row:
            ws2_2[cell.coordinate].value = cell.value

    ws2_2.column_dimensions['A'].width = 25
    ws2_2.column_dimensions['B'].width = 25
    ws2_2.column_dimensions['C'].width = 10
    ws2_2.column_dimensions['D'].width = 30
    ws2_2.column_dimensions['E'].width = 15
    ws2_2.column_dimensions['F'].width = 15
    ws2_2.column_dimensions['G'].width = 20


    # Sheet 3 -------
    ws2_3 = wb2.create_sheet(ws1_3.title)
    for row in ws1_3:
        for cell in row:
            ws2_3[cell.coordinate].value = cell.value

    ws2_3.column_dimensions['A'].width = 30
    ws2_3.column_dimensions['B'].width = 10


    # Sheet 4 -------
    ws2_4 = wb2.create_sheet(ws1_4.title)
    for row in ws1_4:
        for cell in row:
            ws2_4[cell.coordinate].value = cell.value

    ws2_4.column_dimensions['A'].width = 25
    ws2_4.column_dimensions['B'].width = 10

    # Sheet 5 -------
    ws2_5 = wb2.create_sheet(ws1_5.title)
    for row in ws1_5:
        for cell in row:
            ws2_5[cell.coordinate].value = cell.value

    ws2_5.column_dimensions['A'].width = 25
    ws2_5.column_dimensions['B'].width = 25
    ws2_5.column_dimensions['C'].width = 10
    ws2_5.column_dimensions['D'].width = 30
    ws2_5.column_dimensions['E'].width = 15
    ws2_5.column_dimensions['F'].width = 15
    ws2_5.column_dimensions['G'].width = 20


    # Sheet 6 -------
    ws2_6 = wb2.create_sheet(ws1_6.title)
    for row in ws1_6:
        for cell in row:
            ws2_6[cell.coordinate].value = cell.value

    ws2_6.column_dimensions['A'].width = 25
    ws2_6.column_dimensions['B'].width = 10

    # Sheet 7 -------
    ws2_7 = wb2.create_sheet(ws1_7.title)
    for row in ws1_7:
        for cell in row:
            ws2_7[cell.coordinate].value = cell.value

    ws2_7.column_dimensions['A'].width = 25
    ws2_7.column_dimensions['B'].width = 25
    ws2_7.column_dimensions['C'].width = 10
    ws2_7.column_dimensions['D'].width = 30
    ws2_7.column_dimensions['E'].width = 15
    ws2_7.column_dimensions['F'].width = 15
    ws2_7.column_dimensions['G'].width = 20

    # Sheet 8 -------
    ws2_8 = wb2.create_sheet(ws1_8.title)
    for row in ws1_8:
        for cell in row:
            ws2_8[cell.coordinate].value = cell.value

    ws2_8.column_dimensions['A'].width = 25
    ws2_8.column_dimensions['B'].width = 10

    # Sheet 9 -------
    ws2_9 = wb2.create_sheet(ws1_9.title)
    for row in ws1_9:
        for cell in row:
            ws2_9[cell.coordinate].value = cell.value

    ws2_9.column_dimensions['A'].width = 30
    ws2_9.column_dimensions['B'].width = 30
    ws2_9.column_dimensions['C'].width = 30
    ws2_9.column_dimensions['D'].width = 30
    ws2_9.column_dimensions['E'].width = 30
    ws2_9.column_dimensions['F'].width = 30
    ws2_9.column_dimensions['G'].width = 30
    ws2_9.column_dimensions['H'].width = 30
    ws2_9.column_dimensions['I'].width = 30

    # Sheet 10 -------
    ws2_10 = wb2.create_sheet(ws1_10.title)
    for row in ws1_10:
        for cell in row:
            ws2_10[cell.coordinate].value = cell.value

    ws2_10.column_dimensions['A'].width = 30
    ws2_10.column_dimensions['B'].width = 30
    ws2_10.column_dimensions['C'].width = 30
    ws2_10.column_dimensions['D'].width = 30
    ws2_10.column_dimensions['E'].width = 30
    ws2_10.column_dimensions['F'].width = 30
    ws2_10.column_dimensions['G'].width = 30
    ws2_10.column_dimensions['H'].width = 30

    # Sheet 11 -------
    ws2_11 = wb2.create_sheet(ws1_11.title)
    for row in ws1_11:
        for cell in row:
            ws2_11[cell.coordinate].value = cell.value

    ws2_11.column_dimensions['A'].width = 25
    ws2_11.column_dimensions['B'].width = 10

    # Sheet 12 -------
    ws2_12 = wb2.create_sheet(ws1_12.title)
    for row in ws1_12:
        for cell in row:
            ws2_12[cell.coordinate].value = cell.value

    ws2_12.column_dimensions['A'].width = 25
    ws2_12.column_dimensions['B'].width = 10


    # Sheet 13 -------
    ws2_13 = wb2.create_sheet()

    ws2_13.column_dimensions['A'].width = 50
    ws2_13.column_dimensions['B'].width = 50
    ws2_13.column_dimensions['C'].width = 50
    ws2_13.column_dimensions['D'].width = 120

    ws2_13.cell(row = 1 , column = 1).value = "Number of Tweets which their states Defined by Human"
    ws2_13.cell(row = 1 , column = 2).value = "Number of Tweets which their states Defined by Machine"
    ws2_13.cell(row = 1 , column = 3).value = "Accuracy of Tweets' State which Defined by Machine"
    ws2_13.cell(row = 1 , column = 4).value = "Accuracy of Tweets' State for whole Dataset in average, by assuming the accuracy for Tweets' State which Defined by Human is 100% "


    ws2_13.cell(row = 2 , column = 1).value = "All Tweets' States Defined by Human"
    ws2_13.cell(row = 2 , column = 2).value = "0"
    ws2_13.cell(row = 2 , column = 3).value = "No Machine Used In This Dataset"
    ws2_13.cell(row = 2 , column = 4).value = "%100"
                
    # removing main sheet (sheet 0) #because main sheet is empty, by the above coding the data pouring to excel starts at sheet1 not sheet
    std = wb2['Sheet']
    wb2.remove(std)

    
    wb2.save(my_path_output + 'DATASET.xlsx')


def main ():

    import xlrd

    my_path_input = 'D:\\Papers\\Social Network Mining\\Creating_Dataset\\Step6-1\\False Rumors\\False_Rumor_22\\Input\\'
    my_path_output = 'D:\\Papers\\Social Network Mining\\Creating_Dataset\\Step6-1\\False Rumors\\False_Rumor_22\\Output\\'  #address for saving result files (this line need to be re write for new datasets according to the desired address we want to put the results in them)

    workbook = xlrd.open_workbook(my_path_input + 'excel_results.xlsx')
    worksheet1_main = workbook.sheet_by_name('Sheet1')
    
    workbook_whole = xlrd.open_workbook(my_path_input + 'excel_judge_whole.xlsx')
    worksheet1_whole = workbook_whole.sheet_by_name('Sheet1')

    x = check_if_dataset_by_judge_is_correct (worksheet1_whole)
    if x == "Dataset is Ok" :
        dic_retweet_frequency_state = make_dic_retweet_frequency_state (worksheet1_whole)
        Frequency_List,State_List = make_list_of_states_frequencies (worksheet1_main,dic_retweet_frequency_state)
        make_final_dataset (State_List,my_path_input,my_path_output,Frequency_List)
    


main ()
input ("please press enter to exit ...")
