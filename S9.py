#-------------------------------------------------------------
# This program gives us the final excel tabel of dataset named DATASET with all data
#-------------------------------------------------------------
def get_X_and_get_y (worksheet1,worksheet2):

    #get X
    Tweet_ID = worksheet1.col_values(0)
    num_row = len(Tweet_ID)
    labels_tabels = worksheet1.row_values(0)
    num_col = len(labels_tabels)
    row_n = 1
    X = []
    row_data_int = []
    while row_n < num_row :
        row_data = worksheet1.row_values(row_n)
        del row_data [(num_col - 1)] #deleting last element which is Tweet status
        del row_data [0] #deleting first element which is Tweet ID
        for element in row_data :
            row_data_int.append (int (element))
        X.append (row_data_int)
        row_data_int = []
        row_n += 1


    #get y (targets)
    y = worksheet1.col_values(num_col - 1)
    del y [0]

    del Tweet_ID[0]

    features = worksheet1.row_values(0)
    del features[num_col - 1] #deleting status
    del features[0] #deleting first label which is "tweet id" and not a feature


    
    number_tweets_which_their_states_defined_by_human = 0
    freq = worksheet2.col_values(1)
    del freq [0]
    for n in freq :
        number_tweets_which_their_states_defined_by_human += n
    
        
            
    return X,y,Tweet_ID,features,number_tweets_which_their_states_defined_by_human    

#----------------------------------
def machine_learning (X,y) :
    import sklearn
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.model_selection import train_test_split, cross_val_score, cross_val_predict
    from sklearn import metrics
    
    # with cross validation main
    clf_cv = RandomForestClassifier(n_estimators = 1000 , max_depth=2500, random_state=0)
    model_cv = clf_cv.fit(X, y)
    scores = cross_val_score(model_cv, X, y, cv=20) #divide dataset to 20 parts and each time learn with 19 parts and test on one part
    return model_cv,scores

#----------------------------------
def finding_state_of_each_tweet (worksheet1_main,row_n_main,model_cv,Tweet_ID,features,y) :
    # first check to see if the tweet is a retweet or non retweet, because if it is a retweet then we want the retweet id as tweet id
    retweet_id = worksheet1_main.cell(row_n_main,23).value
    if retweet_id != "Not a Retweet" :
        tweet_id = worksheet1_main.cell(row_n_main,23).value
    else :
        tweet_id = worksheet1_main.cell(row_n_main,21).value

    if tweet_id not in Tweet_ID :
        uni_bi_current_tweet = get_uni_bi_grams_of_current_tweet (row_n_main, worksheet1_main)
        tweet_feature_values = []
        ready_for_predict = []
        for feature in features :
            if feature in uni_bi_current_tweet :
                tweet_feature_values.append ("1")
            else :
                tweet_feature_values.append ("0")
        ready_for_predict.append (tweet_feature_values)
        state_current_tweet_list = model_cv.predict (ready_for_predict)
        state_current_tweet = state_current_tweet_list[0]

    else :
        row_number = len (Tweet_ID)
        rn = 0
        while rn < row_number :
            if Tweet_ID[rn] == tweet_id :
                state_current_tweet = y[rn]

            rn += 1


    return state_current_tweet
        
#----------------------------------       
def get_uni_bi_grams_of_current_tweet (row_n_main, worksheet1_main):

    import nltk
    from nltk.tokenize import word_tokenize
    from nltk.corpus import stopwords 
    from nltk.stem import WordNetLemmatizer
    lemmatizer = WordNetLemmatizer()
    from nltk import pos_tag
    import string
    token_lem = []
    tweet_words = []
    tweet_biwords = []
    
    retweet_text = worksheet1_main.cell(row_n_main,22).value
    if retweet_text != "Not a Retweet" :
        text = retweet_text
    else :
        text = worksheet1_main.cell(row_n_main,20).value

    for i,j in pos_tag(word_tokenize(text)):
        if j[0].lower() in ['a','n','v'] : #(a = adjective , n = noun , v = verb)
            token_lem.append(lemmatizer.lemmatize(i,j[0].lower()))
        else :
            token_lem.append(lemmatizer.lemmatize(i))


    stoplist = set(stopwords.words()) #not defining "english" so it can be able to delete all stopwords from all languages ...
    additional = ["'s","?","n't","'i",'’','«','could',"'m",'”','would',"'you","...",'–','‘',"'ll","'re","'no","'re",'--',"'yes","'ve","'it",'``',"''",'http',"'that","'it","'the","'they","'he","'d","mrs","mr","'we"]    
    for word in token_lem :
        if (word.lower() not in stoplist ) and (word not in string.punctuation) and (word.lower() not in additional)  :
            tweet_words.append (word) # this is for function of enter_data_into_dataset
    tweet_words = set (tweet_words)

    for bigram in nltk.bigrams(tweet_words) :
        tweet_biwords.append (bigram)

    uni_bi_current_tweet = []
        
    for unigrams in tweet_words :
        uni_bi_current_tweet.append (unigrams)

    for bigrams in tweet_biwords :
        uni_bi_current_tweet.append (bigrams)
            

    return uni_bi_current_tweet        

#----------------------------------
def make_final_dataset (state_all_tweets,my_path_input,my_path_output,Frequency,scores,number_tweets_which_their_states_defined_by_human) : #it is same as

    import openpyxl as xl

    wb1 = xl.load_workbook(my_path_input + 'excel_results.xlsx')
    
    ws1_1 = wb1.worksheets[0]
    ws1_2 = wb1.worksheets[1]
    ws1_3 = wb1.worksheets[2]
    ws1_4 = wb1.worksheets[3]
    ws1_5 = wb1.worksheets[4]
    ws1_6 = wb1.worksheets[5]
    ws1_7 = wb1.worksheets[6]
    ws1_8 = wb1.worksheets[7]
    ws1_9 = wb1.worksheets[8]
    ws1_10 = wb1.worksheets[9]
    ws1_11 = wb1.worksheets[10]
    ws1_12 = wb1.worksheets[11]

    
    wb2 = xl.Workbook()

    # Sheet 1 -------
    ws2_1 = wb2.create_sheet(ws1_1.title)
    for row in ws1_1:
        for cell in row:
            ws2_1[cell.coordinate].value = cell.value


    #writing the frequency of each tweet in column 27
    ws2_1.cell(row = 1 ,column = 28).value = 'Frequency of Tweet Occurance' #for excel 'row = 1' is same as index = 0 for python , ie the label row
    v = 2
    for fr in Frequency :
        ws2_1.cell(row = v ,column = 28).value = fr
        v += 1
    
    
    #writing state of each tweet in column 28
    ws2_1.cell(row = 1 ,column = 29).value = 'State of Tweet' #for excel 'row = 1' is same as index = 0 for python , ie the label row
    v = 2
    for state in state_all_tweets :
        ws2_1.cell(row = v ,column = 29).value = state
        v += 1

    ws2_1.column_dimensions['A'].width = 25
    ws2_1.column_dimensions['B'].width = 25
    ws2_1.column_dimensions['C'].width = 200
    ws2_1.column_dimensions['D'].width = 15
    ws2_1.column_dimensions['E'].width = 25
    ws2_1.column_dimensions['F'].width = 10
    ws2_1.column_dimensions['G'].width = 30
    ws2_1.column_dimensions['H'].width = 35
    ws2_1.column_dimensions['I'].width = 15
    ws2_1.column_dimensions['J'].width = 15
    ws2_1.column_dimensions['K'].width = 25
    ws2_1.column_dimensions['L'].width = 15
    ws2_1.column_dimensions['M'].width = 25
    ws2_1.column_dimensions['N'].width = 15
    ws2_1.column_dimensions['O'].width = 20
    ws2_1.column_dimensions['P'].width = 20
    ws2_1.column_dimensions['Q'].width = 20
    ws2_1.column_dimensions['R'].width = 15
    ws2_1.column_dimensions['S'].width = 50
    ws2_1.column_dimensions['T'].width = 50
    ws2_1.column_dimensions['U'].width = 200
    ws2_1.column_dimensions['V'].width = 25
    ws2_1.column_dimensions['W'].width = 200
    ws2_1.column_dimensions['X'].width = 25
    ws2_1.column_dimensions['Y'].width = 200
    ws2_1.column_dimensions['Z'].width = 25
    ws2_1.column_dimensions['AA'].width = 25
    ws2_1.column_dimensions['AB'].width = 30
    ws2_1.column_dimensions['AC'].width = 15


    
    
    # Sheet 2 -------
    ws2_2 = wb2.create_sheet(ws1_2.title)
    for row in ws1_2:
        for cell in row:
            ws2_2[cell.coordinate].value = cell.value

    ws2_2.column_dimensions['A'].width = 25
    ws2_2.column_dimensions['B'].width = 25
    ws2_2.column_dimensions['C'].width = 10
    ws2_2.column_dimensions['D'].width = 30
    ws2_2.column_dimensions['E'].width = 15
    ws2_2.column_dimensions['F'].width = 15
    ws2_2.column_dimensions['G'].width = 20


    # Sheet 3 -------
    ws2_3 = wb2.create_sheet(ws1_3.title)
    for row in ws1_3:
        for cell in row:
            ws2_3[cell.coordinate].value = cell.value

    ws2_3.column_dimensions['A'].width = 30
    ws2_3.column_dimensions['B'].width = 10


    # Sheet 4 -------
    ws2_4 = wb2.create_sheet(ws1_4.title)
    for row in ws1_4:
        for cell in row:
            ws2_4[cell.coordinate].value = cell.value

    ws2_4.column_dimensions['A'].width = 25
    ws2_4.column_dimensions['B'].width = 10

    # Sheet 5 -------
    ws2_5 = wb2.create_sheet(ws1_5.title)
    for row in ws1_5:
        for cell in row:
            ws2_5[cell.coordinate].value = cell.value

    ws2_5.column_dimensions['A'].width = 25
    ws2_5.column_dimensions['B'].width = 25
    ws2_5.column_dimensions['C'].width = 10
    ws2_5.column_dimensions['D'].width = 30
    ws2_5.column_dimensions['E'].width = 15
    ws2_5.column_dimensions['F'].width = 15
    ws2_5.column_dimensions['G'].width = 20


    # Sheet 6 -------
    ws2_6 = wb2.create_sheet(ws1_6.title)
    for row in ws1_6:
        for cell in row:
            ws2_6[cell.coordinate].value = cell.value

    ws2_6.column_dimensions['A'].width = 25
    ws2_6.column_dimensions['B'].width = 10

    # Sheet 7 -------
    ws2_7 = wb2.create_sheet(ws1_7.title)
    for row in ws1_7:
        for cell in row:
            ws2_7[cell.coordinate].value = cell.value

    ws2_7.column_dimensions['A'].width = 25
    ws2_7.column_dimensions['B'].width = 25
    ws2_7.column_dimensions['C'].width = 10
    ws2_7.column_dimensions['D'].width = 30
    ws2_7.column_dimensions['E'].width = 15
    ws2_7.column_dimensions['F'].width = 15
    ws2_7.column_dimensions['G'].width = 20

    # Sheet 8 -------
    ws2_8 = wb2.create_sheet(ws1_8.title)
    for row in ws1_8:
        for cell in row:
            ws2_8[cell.coordinate].value = cell.value

    ws2_8.column_dimensions['A'].width = 25
    ws2_8.column_dimensions['B'].width = 10

    # Sheet 9 -------
    ws2_9 = wb2.create_sheet(ws1_9.title)
    for row in ws1_9:
        for cell in row:
            ws2_9[cell.coordinate].value = cell.value

    ws2_9.column_dimensions['A'].width = 30
    ws2_9.column_dimensions['B'].width = 30
    ws2_9.column_dimensions['C'].width = 30
    ws2_9.column_dimensions['D'].width = 30
    ws2_9.column_dimensions['E'].width = 30
    ws2_9.column_dimensions['F'].width = 30
    ws2_9.column_dimensions['G'].width = 30
    ws2_9.column_dimensions['H'].width = 30
    ws2_9.column_dimensions['I'].width = 30

    # Sheet 10 -------
    ws2_10 = wb2.create_sheet(ws1_10.title)
    for row in ws1_10:
        for cell in row:
            ws2_10[cell.coordinate].value = cell.value

    ws2_10.column_dimensions['A'].width = 30
    ws2_10.column_dimensions['B'].width = 30
    ws2_10.column_dimensions['C'].width = 30
    ws2_10.column_dimensions['D'].width = 30
    ws2_10.column_dimensions['E'].width = 30
    ws2_10.column_dimensions['F'].width = 30
    ws2_10.column_dimensions['G'].width = 30
    ws2_10.column_dimensions['H'].width = 30

    # Sheet 11 -------
    ws2_11 = wb2.create_sheet(ws1_11.title)
    for row in ws1_11:
        for cell in row:
            ws2_11[cell.coordinate].value = cell.value

    ws2_11.column_dimensions['A'].width = 25
    ws2_11.column_dimensions['B'].width = 10

    # Sheet 12 -------
    ws2_12 = wb2.create_sheet(ws1_12.title)
    for row in ws1_12:
        for cell in row:
            ws2_12[cell.coordinate].value = cell.value

    ws2_12.column_dimensions['A'].width = 25
    ws2_12.column_dimensions['B'].width = 10


    # Sheet 13 -------
    ws2_13 = wb2.create_sheet()

    ws2_13.column_dimensions['A'].width = 50
    ws2_13.column_dimensions['B'].width = 50
    ws2_13.column_dimensions['C'].width = 50
    ws2_13.column_dimensions['D'].width = 120

    ws2_13.cell(row = 1 , column = 1).value = "Number of Tweets which their states Defined by Human"
    ws2_13.cell(row = 1 , column = 2).value = "Number of Tweets which their states Defined by Machine"
    ws2_13.cell(row = 1 , column = 3).value = "Accuracy of Tweets' State which Defined by Machine"
    ws2_13.cell(row = 1 , column = 4).value = "Accuracy of Tweets' State for whole Dataset in average, by assuming the accuracy for Tweets' State which Defined by Human is 100% "

    tweet_number_all = len(Frequency)
    number_tweets_which_their_states_defined_by_machine = tweet_number_all - number_tweets_which_their_states_defined_by_human 
    accuracy = "%0.2f (+/- %0.2f)" % (scores.mean(), scores.std() * 2)
    average_accuracy_raw = ((number_tweets_which_their_states_defined_by_human) + (number_tweets_which_their_states_defined_by_machine * (scores.mean())))/((number_tweets_which_their_states_defined_by_human + number_tweets_which_their_states_defined_by_machine))*100
    average_accuracy = "%" + str(round (average_accuracy_raw))

    ws2_13.cell(row = 2 , column = 1).value = str (number_tweets_which_their_states_defined_by_human)
    ws2_13.cell(row = 2 , column = 2).value = str (number_tweets_which_their_states_defined_by_machine)
    ws2_13.cell(row = 2 , column = 3).value = accuracy
    ws2_13.cell(row = 2 , column = 4).value = average_accuracy
                
    # removing main sheet (sheet 0) #because main sheet is empty, by the above coding the data pouring to excel starts at sheet1 not sheet
    std = wb2['Sheet']
    wb2.remove(std)

    
    wb2.save(my_path_output + 'DATASET.xlsx')
    
    
#----------------------------------
##  making a dic of retweet and frequencies ----
def make_dic_retweet_frequency (worksheet1_whole):
    Frequency = worksheet1_whole.col_values(1)
    Retweet_ID = worksheet1_whole.col_values(2)
    del Frequency[0]
    del Retweet_ID[0]
    number_tweets = len (Frequency)
        
    dic_retweet_frequency = {}
    k = 0
    while k < number_tweets :
        if Retweet_ID[k] != 'Not a Retweet' :
            dic_retweet_frequency[Retweet_ID[k]]= Frequency[k]
        k += 1

    return dic_retweet_frequency

#----------------------------------
def make_list_of_retweet_frequency (worksheet1_whole,worksheet1_main):
    dic_retweet_frequency = make_dic_retweet_frequency (worksheet1_whole)
    Tweet_ID_main = worksheet1_main.col_values(21)
    Retweet_ID_main = worksheet1_main.col_values(23)

    del Tweet_ID_main[0]
    del Retweet_ID_main[0]
    
    tweet_number = len (Tweet_ID_main)
    Frequency = []
    k = 0
    hh = 0
    while k < tweet_number :
        if Retweet_ID_main[k] != "Not a Retweet" :
            if Retweet_ID_main[k] in dic_retweet_frequency :
                Frequency.append (str(dic_retweet_frequency[Retweet_ID_main[k]]))
            
        else :
            if Tweet_ID_main[k] in dic_retweet_frequency :
                statement = "This Tweet is Origin of " + str(dic_retweet_frequency[Tweet_ID_main[k]]) + " Tweets"
                Frequency.append (statement)
            else :
                Frequency.append ('1')

        k += 1


    return Frequency,tweet_number
    
#----------------------------------
def main ():
    
    import xlrd
    #Getting uni bi grams ----
    my_path_input = 'D:\\Papers\\Social Network Mining\\Creating_Dataset\\Step9\\False Rumors\\False_Rumor_24\\Input\\'
    my_path_output = 'D:\\Papers\\Social Network Mining\\Creating_Dataset\\Step9\\False Rumors\\False_Rumor_24\\Output\\'  #address for saving result files (this line need to be re write for new datasets according to the desired address we want to put the results in them)

    workbook = xlrd.open_workbook(my_path_input + 'excel_learning_dataset.xlsx')
    worksheet1 = workbook.sheet_by_name('Sheet1')
    worksheet2 = workbook.sheet_by_name('Sheet2')

    workbook = xlrd.open_workbook(my_path_input + 'excel_judge_whole.xlsx')
    worksheet1_whole = workbook.sheet_by_name('Sheet1')

    workbook_main = xlrd.open_workbook(my_path_input + 'excel_results.xlsx')
    worksheet1_main = workbook_main.sheet_by_name('Sheet1')

    X,y,Tweet_ID,features,number_tweets_which_their_states_defined_by_human = get_X_and_get_y (worksheet1,worksheet2)
    model_cv,scores = machine_learning (X,y)

    all_rows_main = worksheet1_main.col_values(0)
    number_of_all_rows_in_main_dataset = len (all_rows_main)
    row_n_main = 1
    state_all_tweets = []
    while row_n_main < number_of_all_rows_in_main_dataset : #number_of_all_rows_in_main_dataset  
        state_current_tweet = finding_state_of_each_tweet (worksheet1_main,row_n_main,model_cv,Tweet_ID,features,y)
        state_all_tweets.append (state_current_tweet)
        row_n_main += 1

    dic_retweet_frequency = make_dic_retweet_frequency (worksheet1)
    Frequency,tweet_number = make_list_of_retweet_frequency (worksheet1_whole,worksheet1_main)
    make_final_dataset (state_all_tweets,my_path_input,my_path_output,Frequency,scores,number_tweets_which_their_states_defined_by_human)    

#----------------------------------    

main()
input ("press enter key to exit ...")
